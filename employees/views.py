# Standard Library Imports
import json
from datetime import datetime, timedelta
from io import BytesIO

# Django Core Imports
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import (
    Count, Q, Avg, Sum, F, ExpressionWrapper, 
    DurationField, Case, When, IntegerField
)
from django import forms

# Third-Party Imports
from openpyxl import Workbook
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
import base64

# Local Imports
from .models import (
    Employee, Attendance, SalaryPayment, 
    Department, EmployeePrediction, PerformanceReview
)
from .ai_analytics import TurnoverPredictor, update_all_predictions


class EmployeeForm(forms.ModelForm):
    class Meta:
        model = Employee
        fields = '__all__'
        widgets = {
            'hire_date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'phone': forms.TextInput(attrs={
                'placeholder': 'ادخل رقم الهاتف',
                'class': 'form-control'
            }),
            'image': forms.FileInput(attrs={
                'class': 'form-control',
                'accept': 'image/*'
            })
        }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['manager'].queryset = Employee.objects.annotate(
            num_subordinates=Count('subordinates')
        ).filter(num_subordinates__gt=0)
        self.fields['department'].queryset = Department.objects.all()

    def clean_image(self):
        image = self.cleaned_data.get('image')
        if image and isinstance(image, bytes):
            return ContentFile(image, name='employee_image.jpg')
        return image


class SalaryCalculationForm(forms.Form):
    MONTH_CHOICES = [
        (1, 'يناير'), (2, 'فبراير'), (3, 'مارس'),
        (4, 'أبريل'), (5, 'مايو'), (6, 'يونيو'),
        (7, 'يوليو'), (8, 'أغسطس'), (9, 'سبتمبر'),
        (10, 'أكتوبر'), (11, 'نوفمبر'), (12, 'ديسمبر')
    ]
    
    month = forms.ChoiceField(choices=MONTH_CHOICES, label='الشهر')
    year = forms.IntegerField(
        min_value=2000,
        max_value=datetime.now().year,
        label='السنة'
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        current_date = timezone.now()
        self.fields['month'].initial = current_date.month
        self.fields['year'].initial = current_date.year


# Employee Views
@login_required
def employee_list(request):
    employees = Employee.objects.all().select_related('department', 'manager')
    managers = Employee.objects.annotate(
        num_subordinates=Count('subordinates')
    ).filter(num_subordinates__gt=0)
    
    return render(request, 'employees/employee_list.html', {
        'employees': employees,
        'managers': managers,
        'page_title': 'قائمة الموظفين'
    })


@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(
        Employee.objects.select_related('department', 'manager', 'user'),
        pk=pk
    )
    
    current_date = timezone.now()
    start_date = current_date.replace(day=1)
    end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Attendance Data
    attendances = Attendance.objects.filter(
        employee=employee,
        date__lte=current_date
    ).order_by('-date')[:30]
    
    for attendance in attendances:
        attendance.working_hours_value = attendance.get_working_hours()
    
    # Performance Reviews
    performance_reviews = PerformanceReview.objects.filter(
        employee=employee
    ).order_by('-review_date')[:5]
    
    # Attendance Statistics
    current_month_stats = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    ).aggregate(
        working_days=Count('id', filter=Q(status='present')),
        total_hours=Sum('working_hours')
    )
    
    working_days = current_month_stats['working_days'] or 0
    total_hours = current_month_stats['total_hours'] or 0.0
    attendance_percentage = round((working_days / end_date.day) * 100) if end_date.day > 0 else 0
    
    # Chart Data
    attendance_by_day = Attendance.objects.filter(
        employee=employee,
        date__gte=current_date - timedelta(days=30)
    ).values('date__week_day').annotate(
        count=Count('id', filter=Q(status='present'))
    ).order_by('date__week_day')
    
    # Performance Data
    performance_avg = employee.performance_reviews.aggregate(
        avg_rating=Avg('rating')
    )['avg_rating'] or 0
    performance_avg = round(performance_avg, 2)
    
    # Predictions
    predictions = EmployeePrediction.objects.filter(
        employee=employee
    ).order_by('-prediction_date')[:5]
    
    context = {
        'employee': employee,
        'attendances': attendances,
        'performance_reviews': performance_reviews,
        'working_days': working_days,
        'total_hours': round(total_hours, 2),
        'attendance_percentage': attendance_percentage,
        'daily_salary': round(employee.salary / 30, 2),
        'employee_image_url': employee.get_image_url(),
        'current_month': current_date.month,
        'current_year': current_date.year,
        'predictions': predictions,
        'performance_avg': performance_avg,
        'turnover_risk': predictions[0].turnover_risk if predictions else 0,
        'attendance_data': json.dumps({
            'labels': ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت'],
            'data': [day['count'] for day in attendance_by_day]
        }, cls=DjangoJSONEncoder),
        'performance_data': json.dumps({
            'labels': [review.review_date.strftime('%Y-%m-%d') for review in performance_reviews],
            'data': [review.rating for review in performance_reviews]
        }, cls=DjangoJSONEncoder)
    }
    
    return render(request, 'employees/employee_detail.html', context)


# Analytics Views
@login_required
@permission_required('employees.view_analytics', raise_exception=True)
def analytics_dashboard(request):
    if request.method == 'POST' and 'run_predictions' in request.POST:
        if update_all_predictions():
            messages.success(request, 'تم تشغيل التحليل بنجاح')
        else:
            messages.warning(request, 'لا توجد بيانات كافية للتحليل')

    high_risk = EmployeePrediction.objects.filter(
        turnover_risk__gte=0.5
    ).select_related('employee').order_by('-turnover_risk')[:10]

    # Department performance analytics
    department_stats = Department.objects.annotate(
        avg_performance=Avg('employee__performance_reviews__rating'),
        employee_count=Count('employee')
    ).filter(employee_count__gt=0)

    return render(request, 'employees/analytics.html', {
        'high_risk_employees': high_risk,
        'department_stats': department_stats,
        'plot_url': create_analytics_plots(),
        'current_date': timezone.now().strftime("%Y-%m-%d")
    })


def create_analytics_plots():
    try:
        # Performance Distribution
        plt.figure(figsize=(12, 6))
        
        # Performance by Department
        department_stats = Department.objects.annotate(
            avg_performance=Avg('employee__performance_reviews__rating'),
            employee_count=Count('employee')
        ).filter(employee_count__gt=0)
        
        if department_stats:
            plt.subplot(1, 2, 1)
            plt.bar(
                [dept.name for dept in department_stats],
                [dept.avg_performance for dept in department_stats]
            )
            plt.title('متوسط الأداء حسب القسم')
            plt.xticks(rotation=45)
            
            # Turnover Risk Distribution
            plt.subplot(1, 2, 2)
            risk_dist = EmployeePrediction.objects.values('turnover_risk').annotate(
                count=Count('id')
            ).order_by('turnover_risk')
            
            if risk_dist:
                plt.hist(
                    [x['turnover_risk'] for x in risk_dist],
                    bins=10,
                    color='skyblue',
                    edgecolor='black'
                )
                plt.title('توزيع مخاطر الاستقالة')
        
        buffer = BytesIO()
        plt.tight_layout()
        plt.savefig(buffer, format='png', bbox_inches='tight')
        plt.close()
        
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    except Exception as e:
        print(f"Error generating plots: {e}")
        return None


# Performance Review Views
@login_required
@permission_required('employees.manage_reviews', raise_exception=True)
def add_performance_review(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    
    if request.method == 'POST':
        reviewer = request.user
        rating = request.POST.get('rating')
        comments = request.POST.get('comments', '')
        
        if not rating:
            messages.error(request, 'يجب تحديد تقييم الأداء')
        else:
            PerformanceReview.objects.create(
                employee=employee,
                reviewer=reviewer,
                rating=rating,
                comments=comments,
                review_date=timezone.now()
            )
            messages.success(request, 'تم إضافة تقييم الأداء بنجاح')
            return redirect('employee_detail', pk=employee_id)
    
    return render(request, 'employees/add_review.html', {
        'employee': employee,
        'rating_choices': PerformanceReview.RATING_CHOICES
    })


@login_required
def performance_review_detail(request, review_id):
    review = get_object_or_404(
        PerformanceReview.objects.select_related('employee', 'reviewer'),
        pk=review_id
    )
    
    if not request.user.has_perm('employees.manage_reviews') and request.user != review.reviewer:
        return HttpResponseForbidden("ليس لديك صلاحية لعرض هذه الصفحة")
    
    return render(request, 'employees/review_detail.html', {
        'review': review
    })


# Salary Views
@login_required
@permission_required('employees.calculate_salary', raise_exception=True)
def calculate_salary(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        form = SalaryCalculationForm(request.POST)
        if form.is_valid():
            month = int(form.cleaned_data['month'])
            year = int(form.cleaned_data['year'])
            
            if SalaryPayment.objects.filter(employee=employee, month=month, year=year).exists():
                messages.warning(request, 'تم حساب الراتب لهذا الشهر مسبقاً')
                return redirect('employee_detail', pk=employee.pk)
            
            # Salary Calculation Logic
            start_date = datetime(year, month, 1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            stats = Attendance.objects.filter(
                employee=employee,
                date__range=[start_date, end_date]
            ).aggregate(
                working_days=Count('id', filter=Q(status='present')),
                absent_days=Count('id', filter=Q(status='absent')),
                total_hours=Sum('working_hours')
            )
            
            daily_salary = employee.salary / 30
            base_salary = round(daily_salary * (stats['working_days'] or 0), 2)
            net_salary = base_salary - employee.deductions + employee.bonuses
            
            payment = SalaryPayment(
                employee=employee,
                month=month,
                year=year,
                working_days=stats['working_days'] or 0,
                absent_days=stats['absent_days'] or 0,
                base_salary=base_salary,
                deductions=employee.deductions,
                bonuses=employee.bonuses,
                net_salary=net_salary,
                notes=f"تم الحساب بواسطة {request.user.username}"
            )
            
            if request.user.has_perm('employees.approve_salary'):
                payment.approved_by = getattr(request.user, 'employee_profile', None)
            
            payment.save()
            messages.success(request, 'تم حساب الراتب بنجاح')
            return redirect('employee_detail', pk=employee.pk)
    
    form = SalaryCalculationForm()
    return render(request, 'employees/calculate_salary.html', {
        'employee': employee,
        'form': form,
        'current_month': timezone.now().month,
        'current_year': timezone.now().year,
        'months': SalaryCalculationForm.MONTH_CHOICES,
        'daily_salary': round(employee.salary / 30, 2)
    })


# Export Views
@login_required
@permission_required('employees.export_employees', raise_exception=True)
def export_employees_to_excel(request):
    employees = Employee.objects.all() \
        .select_related('department', 'manager') \
        .order_by('department__name', 'name')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الموظفون"
    
    columns = [
        'الرقم', 'الاسم', 'القسم', 'المدير المسؤول',
        'نوع التوظيف', 'تاريخ التعيين', 'الراتب', 'الهاتف', 'البريد الإلكتروني'
    ]
    ws.append(columns)
    
    for emp in employees:
        ws.append([
            emp.id,
            emp.name,
            emp.department.name if emp.department else 'بدون قسم',
            emp.manager.name if emp.manager else 'بدون مدير',
            emp.get_employee_type_display(),
            emp.hire_date.strftime("%Y-%m-%d") if emp.hire_date else 'غير محدد',
            f"{emp.salary:,.2f}",
            emp.phone or 'غير متوفر',
            emp.email or 'غير متوفر'
        ])
    
    wb.save(response)
    return response


def employee_attendance_report(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    attendances = Attendance.objects.filter(employee=employee).order_by('-date')[:30]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{employee.id}.pdf"'

    p = canvas.Canvas(response)
    
    # PDF Content
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, f"تقرير حضور الموظف: {employee.name}")
    
    p.setFont("Helvetica", 12)
    p.drawString(100, 780, f"القسم: {employee.department.name if employee.department else 'غير محدد'}")
    p.drawString(100, 760, f"الفترة: آخر 30 يوم")

    # Table Headers
    p.setFont("Helvetica-Bold", 12)
    headers = ["التاريخ", "الحالة", "وقت الدخول", "وقت الخروج", "المدة"]
    for i, header in enumerate(headers):
        p.drawString(100 + i*100, 730, header)

    # Table Data
    p.setFont("Helvetica", 10)
    y = 710
    for attendance in attendances:
        p.drawString(100, y, str(attendance.date))
        p.drawString(200, y, attendance.get_status_display())
        p.drawString(300, y, str(attendance.check_in) if attendance.check_in else "-")
        p.drawString(400, y, str(attendance.check_out) if attendance.check_out else "-")
        p.drawString(500, y, f"{attendance.working_hours} ساعات" if attendance.working_hours else "-")
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response


# Analytics Views4encode(buffer.getvalue()).decode('utf-8')


@login_required
@permission_required('employees.export_analytics', raise_exception=True)
def export_analytics(request):
    predictions = EmployeePrediction.objects.all() \
        .select_related('employee') \
        .order_by('-turnover_risk')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="analytics_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "التحليلات"
    
    columns = [
        'الموظف', 'القسم', 'الراتب',
        'احتمال الاستقالة', 'اتجاه الأداء', 'التاريخ', 'التوصية'
    ]
    ws.append(columns)
    
    for pred in predictions:
        ws.append([
            pred.employee.name,
            pred.employee.department.name if pred.employee.department else 'بدون قسم',
            f"{pred.employee.salary:,.2f}",
            f"{pred.turnover_risk:.2%}",
            f"{pred.performance_trend:.2f}",
            pred.prediction_date.strftime("%Y-%m-%d"),  # إصلاح هنا
            pred.recommended_action or "لا توجد توصية"
        ])
    
    # Add statistics sheet if predictor is available
    try:
        predictor = TurnoverPredictor()
        df = predictor.prepare_training_data()
        
        if not df.empty:
            ws2 = wb.create_sheet(title="الإحصائيات")
            stats = df.describe().transpose()
            ws2.append(['المقياس', 'الراتب', 'الأداء', 'الغياب'])
            for index, row in stats.iterrows():
                ws2.append([
                    index,
                    row.get('salary', ''),
                    row.get('performance', ''),
                    row.get('absences', '')
                ])
    except Exception as e:
        print(f"Error generating stats sheet: {e}")
    
    wb.save(response)
    return response



@login_required
@permission_required('employees.run_predictions', raise_exception=True)
def update_predictions(request):
    try:
        success = update_all_predictions()
        if success:
            messages.success(request, 'تم تحديث التنبؤات بنجاح')
        else:
            messages.warning(request, 'لا توجد بيانات كافية لتحديث التنبؤات')
    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء تحديث التنبؤات: {str(e)}')
    
    return redirect('analytics_dashboard')

@login_required
@permission_required('employees.change_employee', raise_exception=True)
def edit_employee(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث بيانات الموظف بنجاح')
            return redirect('employees:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm(instance=employee)
    
    return render(request, 'employees/employee_form.html', {
        'form': form,
        'employee': employee,
        'page_title': f'تعديل موظف: {employee.name}'
    })